from openpyxl import load_workbook
import pandas as pd


class ExcelParser:

    def safe_float(self, value):
        try:
            if value is None:
                return None
            if str(value).strip() == "":
                return None
            return float(value)
        except:
            return None

    def safe_int(self, value):
        try:
            if value is None:
                return None
            if str(value).strip() == "":
                return None
            return int(float(value))
        except:
            return None

    def parse_levels(self, filepath: str):

        wb = load_workbook(filepath, data_only=True)
        sheet = wb.active

        rows = sheet.values
        columns = next(rows)

        columns = [str(c).strip().lower() for c in columns]
        df = pd.DataFrame(rows, columns=columns)

        stock_col = next((c for c in df.columns if "stock" in c), None)
        r1f_col = next((c for c in df.columns if "r1" in c), None)
        r2f_col = next((c for c in df.columns if "r2" in c), None)
        s1f_col = next((c for c in df.columns if "s1" in c), None)
        s2_col = next((c for c in df.columns if c.strip() == "s2"), None)
        qty_col = next((c for c in df.columns if "quantity" in c), None)
        buffer_col = next((c for c in df.columns if "sl" in c), None)

        required = {
            "stock": stock_col,
            "R1,F": r1f_col,
            "R2,F": r2f_col,
            "S1F": s1f_col,
            "S2": s2_col,
            "Quantity": qty_col,
            "SL_BUFFER": buffer_col,
        }

        missing = [k for k, v in required.items() if v is None]
        if missing:
            raise ValueError(f"Missing columns: {missing}")

        result = []

        for _, row in df.iterrows():

            stock = row.get(stock_col)
            if not stock:
                continue

            symbol = str(stock).strip().upper()
            if not symbol.endswith("-EQ"):
                symbol += "-EQ"

            r1f = self.safe_float(row.get(r1f_col))
            r2f = self.safe_float(row.get(r2f_col))
            s1f = self.safe_float(row.get(s1f_col))
            s2 = self.safe_float(row.get(s2_col))
            qty = self.safe_int(row.get(qty_col))
            buffer = self.safe_float(row.get(buffer_col))

            if not qty:
                qty = 1
            if not buffer:
                buffer = 0

            result.append({
                "symbol": symbol,
                "r1f": r1f,
                "r2f": r2f,
                "s1f": s1f,
                "s2": s2,
                "quantity": qty,
                "sl_buffer": buffer
            })

        return result




#////////////////////////////////==================================================
# from openpyxl import load_workbook
# import pandas as pd

# class ExcelParser:

#     def parse_levels(self, filepath: str):

#         wb = load_workbook(filepath, data_only=True)
#         sheet = wb.active

#         rows = sheet.values
#         columns = next(rows)

#         df = pd.DataFrame(rows, columns=[str(c).strip() for c in columns])

#         needed = ["stock", "open price", "R1,F", "R2,F", "S1F", "S2", "Quantity", "SL_BUFFER"]
#         missing = [c for c in needed if c not in df.columns]

#         if missing:
#             raise ValueError(f"Missing columns: {missing}")

#         result = []

#         for _, row in df.iterrows():

#             stock = row.get("stock")
#             if pd.isna(stock):
#                 continue

#             symbol = str(stock).strip().upper()
#             if not symbol.endswith("-EQ"):
#                 symbol += "-EQ"

#             result.append({
#                 "symbol": symbol,
#                 "open_price": float(row["open price"]),
#                 "r1f": float(row["R1,F"]),
#                 "r2f": float(row["R2,F"]),
#                 "s1f": float(row["S1F"]),
#                 "s2": float(row["S2"]),
#                 "quantity": int(row["Quantity"]),
#                 "sl_buffer": float(row["SL_BUFFER"]),
#             })

#         return result






# class ExcelParser:
#     def parse_levels(self, filepath: str):

#         # ✅ data_only=True means calculated values read honge
#         wb = load_workbook(filepath, data_only=True)
#         sheet = wb.active  # first sheet

#         rows = sheet.values
#         columns = next(rows)

#         df = pd.DataFrame(rows, columns=[str(c).strip() for c in columns])

#         needed = ["stock", "Buy", "SELL"]
#         missing = [c for c in needed if c not in df.columns]
#         if missing:
#             raise ValueError(f"Missing columns: {missing}. Required: {needed}")

#         result = []
#         for idx, row in df.iterrows():
#             row_no = idx + 2

#             stock = row.get("stock")
#             if pd.isna(stock) or str(stock).strip() == "":
#                 continue

#             symbol = str(stock).strip().upper()
#             if not symbol.endswith("-EQ"):
#                 symbol += "-EQ"

#             buy = row.get("Buy")
#             sell = row.get("SELL")

#             buy_level = float(buy) if pd.notna(buy) and float(buy) > 0 else None
#             sell_level = float(sell) if pd.notna(sell) and float(sell) > 0 else None

#             result.append({
#                 "row_no": row_no,
#                 "symbol": symbol,
#                 "buy_level": buy_level,
#                 "sell_level": sell_level
#             })

#         return result





# # app/services/excel_parser.py
# from typing import Any, Dict, List, Tuple, Optional
# import pandas as pd


# class ExcelParser:
#     """
#     Supports 2 formats:
#     1) Old strict format: sheet=signals with required columns
#     2) Your custom format: sheet any, columns = stock, Buy, SELL
#     """

#     STRICT_SHEET_NAME = "signals"

#     def parse(self, filepath: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
#         # try strict format first
#         try:
#             df = pd.read_excel(filepath, sheet_name=self.STRICT_SHEET_NAME, engine="openpyxl")
#             df.columns = [str(c).strip() for c in df.columns]
#             return self._parse_strict(df)
#         except Exception:
#             # fallback to custom format
#             df = pd.read_excel(filepath, sheet_name=0, engine="openpyxl")
#             df.columns = [str(c).strip() for c in df.columns]
#             return self._parse_custom(df)

#     # -------------------- STRICT FORMAT --------------------
#     def _parse_strict(self, df: pd.DataFrame):
#         REQUIRED_COLUMNS = [
#             "symbol",
#             "exchange",
#             "action",
#             "quantity",
#             "order_type",
#             "price",
#             "product_type",
#             "validity",
#             "tag",
#         ]
#         missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
#         if missing:
#             raise ValueError(f"Missing required columns in 'signals': {missing}")

#         valid_rows, invalid_rows = [], []

#         for idx, row in df.iterrows():
#             row_no = idx + 2
#             item, err = self._validate_strict_row(row)
#             if err:
#                 invalid_rows.append({
#                     "row_no": row_no,
#                     "tag": str(row.get("tag") or ""),
#                     "symbol": str(row.get("symbol") or ""),
#                     "action": str(row.get("action") or ""),
#                     "status": "FAILED",
#                     "order_id": None,
#                     "remarks": err,
#                 })
#             else:
#                 item["row_no"] = row_no
#                 valid_rows.append(item)

#         return valid_rows, invalid_rows

#     def _validate_strict_row(self, row: pd.Series):
#         def norm(x):
#             if pd.isna(x): return ""
#             return str(x).strip()

#         symbol = norm(row.get("symbol"))
#         exchange = norm(row.get("exchange")).upper()
#         action = norm(row.get("action")).upper()
#         tag = norm(row.get("tag"))
#         order_type = norm(row.get("order_type")).upper()
#         product_type = norm(row.get("product_type")).upper()
#         validity = norm(row.get("validity")).upper()

#         try:
#             quantity = int(row.get("quantity"))
#         except Exception:
#             quantity = 0

#         price = None
#         if not pd.isna(row.get("price")):
#             try:
#                 price = float(row.get("price"))
#             except Exception:
#                 price = None

#         if not tag:
#             return None, "Missing tag"
#         if not symbol:
#             return None, "Missing symbol"
#         if exchange not in ("NSE", "BSE"):
#             return None, "exchange must be NSE/BSE"
#         if action not in ("BUY", "SELL"):
#             return None, "action must be BUY/SELL"
#         if quantity <= 0:
#             return None, "quantity must be >0"
#         if order_type not in ("MARKET", "LIMIT"):
#             return None, "order_type must be MARKET/LIMIT"
#         if order_type == "LIMIT" and (price is None or price <= 0):
#             return None, "LIMIT requires price"
#         if product_type not in ("INTRADAY", "DELIVERY"):
#             return None, "product_type must be INTRADAY/DELIVERY"
#         if validity != "DAY":
#             return None, "validity must be DAY"

#         return {
#             "tag": tag,
#             "symbol": symbol,
#             "exchange": exchange,
#             "action": action,
#             "quantity": quantity,
#             "order_type": order_type,
#             "price": price,
#             "product_type": product_type,
#             "validity": validity,
#             "remarks": norm(row.get("remarks")),
#         }, None

#     # -------------------- CUSTOM FORMAT (YOUR EXCEL) --------------------
#     def _parse_custom(self, df: pd.DataFrame):
#         """
#         Required columns: stock, Buy, SELL
#         We generate 0/1/2 signals per row.
#         """
#         needed = ["stock", "Buy", "SELL"]
#         missing = [c for c in needed if c not in df.columns]
#         if missing:
#             raise ValueError(f"Missing columns: {missing}. Required columns are {needed}")

#         valid_rows: List[Dict[str, Any]] = []
#         invalid_rows: List[Dict[str, Any]] = []

#         for idx, row in df.iterrows():
#             row_no = idx + 2
#             stock = str(row.get("stock") or "").strip()
#             if not stock:
#                 continue

#             # normalize stock -> TRADING SYMBOL
#             symbol = stock.upper()
#             if not symbol.endswith("-EQ"):
#                 symbol = symbol + "-EQ"

#             # BUY signal
#             buy_price = self._to_float(row.get("Buy"))
#             if buy_price and buy_price > 0:
#                 valid_rows.append({
#                     "row_no": row_no,
#                     "tag": f"{symbol}_BUY_{row_no}",
#                     "symbol": symbol,
#                     "exchange": "NSE",
#                     "action": "BUY",
#                     "quantity": 1,
#                     "order_type": "LIMIT",
#                     "price": buy_price,
#                     "product_type": "INTRADAY",
#                     "validity": "DAY",
#                     "remarks": "generated from custom excel",
#                 })

#             # SELL signal
#             sell_price = self._to_float(row.get("SELL"))
#             if sell_price and sell_price > 0:
#                 valid_rows.append({
#                     "row_no": row_no,
#                     "tag": f"{symbol}_SELL_{row_no}",
#                     "symbol": symbol,
#                     "exchange": "NSE",
#                     "action": "SELL",
#                     "quantity": 1,
#                     "order_type": "LIMIT",
#                     "price": sell_price,
#                     "product_type": "INTRADAY",
#                     "validity": "DAY",
#                     "remarks": "generated from custom excel",
#                 })

#         return valid_rows, invalid_rows

#     def _to_float(self, x):
#         if pd.isna(x):
#             return None
#         try:
#             return float(x)
#         except Exception:
#             return None
